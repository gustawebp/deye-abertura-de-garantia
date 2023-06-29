import PySimpleGUI as sg
from openpyxl import Workbook
import os
import random
from datetime import date
from date_today import formatted_date
from input_rules import fault_options, model_options, engineer_options, distributor_options



sg.theme('DefaultNoMoreNagging')
sg.theme('LightBlue2') 

# Definir um estilo personalizado
sg.set_options(font=('Helvetica', 12))  # Def

layout = [

    [sg.Text('CASE NUMBER:')],
    [sg.Input(key='case_number', size=(20,1))],

    [sg.Text('PRIORITY:')],
    [sg.Input(key='priority', size=(20,1))],

    [sg.Text('NAME:')],
    [sg.Input(key='name', size=(20,1))],

#    [sg.Text('STATUS:')],
#   [sg.Input(key='status', size=(20,1))],

    [sg.Text('SELECT AN ENGINEER')],
    [sg.Combo(engineer_options, key='combo_field', default_value='ENGINEER')],

    [sg.Text('REPLACE OR FIX?:')],
    [sg.Input(key='replace_or_fix', size=(20,1))],

    [sg.Text('CASE ORIGIN:')],
    [sg.Input(key='case_origin', size=(20,1))],

    [sg.Text('SELECT A DISTRIBUITOR:')],
    [sg.Combo(distributor_options, key='combo_distribuitor', default_value='ENGINEER')],


   [sg.Text('Selecione um modelo:')],
    [sg.Combo(model_options, key='model_combo', default_value='MODEL')],

    [sg.Text('SN:')],
    [sg.Input(key='sn', size=(20,1))],

    [sg.Text('SELECT A FAULT')],
    [sg.Combo(fault_options, key='fault_combo', default_value='FAULT')],

    [sg.Text('COMMENTS:')],
    [sg.Input(key='comments', size=(20,1))],

    [sg.Button('Ok'), sg.Button('Cancelar')]
]

window = sg.Window('Inserir Dados', layout, size=(350, 680))

while True:
    event, values = window.read()

    if event == sg.WINDOW_CLOSED or event == 'Cancelar':
        break

    if event == 'Ok':
        case_number = values['case_number']
        priority = values['priority']
        name = values['name']
        status = 'WAITING FOR APPROVAL'
        replace_or_fix = values['replace_or_fix']
        case_origin = values['case_origin']
        responsible_engineer = 'GUSTAVO CAMPOS'
        distributor = values['combo_distribuitor']
        model = values['model_combo']
        sn = values['sn']
        fault = values['fault_combo']
        comments = values['comments']
        opening_case_date = formatted_date

        if fault not in fault_options:
            window['error_message'].update('Opção de fault inválida.')
            continue

        # CREATE EXCEL
        workbook = Workbook()
        sheet = workbook.active

        headers = [
            'CASE NUMBER', 'PRIORITY', 'NAME', 'STATUS', 'REPLACE OR FIX?', 'CASE ORIGIN',
            'RESPONSIBLE ENGINEER', 'DISTRIBUTOR', 'MODEL', 'SN', 'FAULT', 'COMMENTS', 'OPENING CASE DATE'
        ]

        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        data_row = [
            case_number, priority, name, status, replace_or_fix, case_origin,
            responsible_engineer, distributor, model, sn, fault, comments, opening_case_date
        ]

        for col_num, value in enumerate(data_row, 1):
            sheet.cell(row=2, column=col_num, value=value)

        # SAVE EXCEL FILE
        file_path = 'dados.xlsx'
        workbook.save(file_path)

        sg.popup(f'Dados salvos em {file_path}')

window.close()
